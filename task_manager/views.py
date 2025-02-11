from datetime import datetime, timedelta, timezone
import re
import dateparser
import pytz
import pandas as pd
from openpyxl.styles import Font, Border, Side
from django.contrib.postgres.aggregates import ArrayAgg
from django.utils.timezone import now
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from django.core.cache import cache
from django.db import IntegrityError
from django.db.models.functions import Concat, TruncDate
from django.db.models import F, Value, Count, Subquery, OuterRef
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from .forms import *
from django.db.models import Q, Sum
from django.utils import timezone


def admin_required(login_url=None):
    return user_passes_test(lambda u: u.is_authenticated and u.is_admin, login_url=login_url)


# Отделы
@admin_required(login_url='login')
def department(request):
    # Попробуем получить отделы из кэша
    departmentObj = cache.get('department_list')
    if not departmentObj:
        # Если кэш пуст, загружаем данные из базы и кэшируем
        departmentObj = Department.objects.annotate(user_count=Count('users'))
        cache.set('department_list', departmentObj, timeout=60 * 15)  # Кэшируем на 15 минут

    selected_departments = request.GET.getlist('department')

    if selected_departments:
        selected_departments = selected_departments[0].split(',')
        departmentObj = departmentObj.filter(title__in=selected_departments)

    form = DepartmentForm()
    return render(request, 'task_manager/departments_list.html', {'departments': departmentObj, 'form': form})


def createDepartment(request):
    form = DepartmentForm()

    if request.method == 'POST':
        form = DepartmentForm(request.POST, request.FILES)
        if form.is_valid():
            department = form.save(commit=False)
            department.save()
            cache.delete('department_list')  # Удаляем кэш после создания
            return redirect('department-list')

    context = {'form': form}
    return render(request, 'task_manager/department_form.html', context)


def updateDepartment(request, pk):
    department = get_object_or_404(Department, pk=pk)
    form = DepartmentForm(instance=department)

    if request.method == 'POST':
        form = DepartmentForm(request.POST, request.FILES, instance=department)
        if form.is_valid():
            department = form.save()
            cache.delete('department_list')  # Удаляем кэш после обновления
            return redirect('department-list')

    context = {'form': form, 'department': department}
    return render(request, 'task_manager/department_form.html', context)


def deleteDepartment(request, pk):
    department = get_object_or_404(Department, pk=pk)

    # Серверная проверка: есть ли пользователи, если да — не удаляем
    if department.users.exists():
        messages.error(request, 'Невозможно удалить отдел, так как к нему привязаны пользователи.')
        return redirect('department-list')

    if request.method == 'POST':
        department.delete()
        cache.delete('department_list')
        return redirect('department-list')

    context = {'object': department}
    return render(request, 'task_manager/department_form.html', context)


# Должности
@admin_required(login_url='login')
def role(request):
    # Пробуем получить список ролей из кэша
    roleObj = cache.get('role_list')

    if not roleObj:
        roleObj = Role.objects.annotate(user_count=Count('users'))
        cache.set('role_list', roleObj, 600)  # Кэшируем на 10 минут (600 секунд)

    selected_roles = request.GET.getlist('role')

    if selected_roles:
        selected_roles = selected_roles[0].split(',')
        roleObj = roleObj.filter(title__in=selected_roles)

    form = RoleForm()
    return render(request, 'task_manager/roles_list.html', {'roles': roleObj, 'form': form})


def createRole(request):
    form = RoleForm()

    if request.method == 'POST':
        form = RoleForm(request.POST, request.FILES)
        if form.is_valid():
            role = form.save(commit=False)
            role.save()

            # Сбрасываем кэш списка ролей
            cache.delete('role_list')

            return redirect('role-list')

    context = {'form': form}
    return render(request, 'task_manager/role_form.html', context)


def updateRole(request, pk):
    role = get_object_or_404(Role, pk=pk)
    form = RoleForm(instance=role)

    if request.method == 'POST':
        form = RoleForm(request.POST, request.FILES, instance=role)
        if form.is_valid():
            role = form.save()

            # Сбрасываем кэш списка ролей
            cache.delete('role_list')

            return redirect('role-list')

    context = {'form': form, 'role': role}
    return render(request, 'task_manager/role_form.html', context)


def deleteRole(request, pk):
    role = get_object_or_404(Role, pk=pk)

    # Серверная проверка: есть ли пользователи, если да — не удаляем
    if role.users.exists():
        messages.error(request, 'Невозможно удалить роль, так как к ней привязаны пользователи.')
        return redirect('role-list')

    if request.method == 'POST':
        role.delete()

        # Сбрасываем кэш списка ролей
        cache.delete('role_list')

        return redirect('role-list')

    context = {'object': role}
    return render(request, 'task_manager/role_form.html', context)


# Юзеры
@admin_required(login_url='login')
def user(request):
    # Кэшируем списки отделов и ролей, чтобы избежать избыточных запросов
    departments = cache.get_or_set('departments_cache', Department.objects.all(), timeout=60 * 15)
    roles = cache.get_or_set('roles_cache', Role.objects.all(), timeout=60 * 15)

    # Фильтры из GET-запроса
    selected_departments = request.GET.getlist('department')
    selected_users = request.GET.getlist('employees')
    selected_roles = request.GET.getlist('role')
    selected_status = request.GET.get('is_admin')
    selected_status_user = request.GET.getlist('status_user')

    # Получаем пользователей с учетом фильтров
    users = CustomUser.objects.select_related('department', 'role')

    status_mapping = {
        'Черновик': 'draft',
        'Приглашен': 'invited',
        'Активен': 'active',
        'Уволен': 'fired'
    }

    # Применяем фильтры, если они указаны
    if selected_status_user:
        selected_status_user = selected_status_user[0].split(',')
        mapped_statuses = [status_mapping.get(status) for status in selected_status_user if status in status_mapping]
        users = users.filter(status__in=mapped_statuses)
    if selected_departments:
        selected_departments = selected_departments[0].split(',')
        users = users.filter(department__title__in=selected_departments)
    if selected_status:
        if selected_status.lower() == 'admin':  # Преобразуем в boolean
            users = users.filter(is_admin=True)
        elif selected_status.lower() == 'user':
            users = users.filter(is_admin=False)
    if selected_roles:
        selected_roles = selected_roles[0].split(',')
        users = users.filter(role__title__in=selected_roles)
    if selected_users:
        user_filters = Q()
        user_list = selected_users[0].split(',')
        for full_name in user_list:
            first_name, last_name = full_name.strip().split(' ', 1)
            user_filters |= Q(first_name__iexact=first_name, last_name__iexact=last_name)
        users = users.filter(user_filters)

    form = CustomUserCreationForm()

    return render(request, 'task_manager/users_list.html', {
        'users': users,
        'form': form,
        'departments': departments,
        'roles': roles,
    })


def createUser(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST, request.FILES)
        if form.is_valid():
            user = form.save()
            # Очищаем кэш списка пользователей и отделов/ролей
            cache.delete('user_list')
            cache.delete('departments_cache')
            cache.delete('roles_cache')

            return redirect('user-list')
    else:
        form = CustomUserCreationForm()

    return render(request, 'task_manager/user_form.html', {'form': form})


def updateUser(request, pk):
    user = get_object_or_404(CustomUser, pk=pk)
    if request.method == 'POST':
        form = CustomUserChangeForm(request.POST, request.FILES, instance=user)
        if form.is_valid():
            form.save()

            # Очищаем кэш после обновления пользователя
            cache.delete('user_list')
            cache.delete('departments_cache')
            cache.delete('roles_cache')

            return redirect('user-list')
    else:
        form = CustomUserChangeForm(instance=user)

    context = {'form': form, 'user': user}
    return render(request, 'task_manager/user_form.html', context)


def deleteUser(request, pk):
    user = get_object_or_404(CustomUser, pk=pk)
    if request.method == 'POST':
        # Устанавливаем статус "уволен"
        user.status = 'fired'

        # Удаляем данные для входа
        user.username = f"fired_{user.id}"  # Уникальное имя пользователя для сохранения истории
        user.set_password(None)  # Удаляем пароль, делая вход невозможным

        user.save()

        # Очищаем кэш после удаления пользователя
        cache.delete('user_list')
        cache.delete('departments_cache')
        cache.delete('roles_cache')

        return redirect('user-list')

    return render(request, 'task_manager/user_confirm_delete.html', {'user': user})


def user_login(request):
    if request.method == 'POST':
        identifier = request.POST['username']
        password = request.POST['password']
        User = get_user_model()
        try:
            user = User.objects.get(email=identifier)
            username = user.username
        except User.DoesNotExist:
            username = identifier  # Если это логин, оставляем как есть

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            # Если пользователь пригашен и входит впервые, меняем статус на активный
            if user.status == 'invited':
                user.status = 'active'
                print(user.status)
                user.save()
                print(user.status)
                # Очищаем кэш после обновления пользователя
                cache.delete('user_list')

            return redirect('user-dashboard')  # Redirect to your desired page after login
        else:
            messages.error(request, 'Invalid username or password')
            return redirect('login')
    return render(request, 'task_manager/login.html')


def user_logout(request):
    if request.user.is_authenticated:
        logout(request)
    return redirect('login')


def send_invitation(request, pk):
    # Выводим для отладки
    # print(f"Received pk from URL: {pk}")
    # print(f"Logged in user pk: {request.user.pk}")
    user = get_object_or_404(CustomUser, pk=pk)
    #print(f'invite: before {user.status}')
    if user.status != 'invited' and user.status != 'active':
        # Генерация случайного логина
        username = get_random_string(length=8)
        while CustomUser.objects.filter(username=username).exists():
            username = get_random_string(length=8)
        user.username = username

        # Генерация случайного пароля
        password = get_random_string(length=8)
        user.set_password(password)
        user.status = 'invited'
        user.save()

        # Очищаем кэш после обновления пользователя
        cache.delete('user_list')

        #print(f'invite: after {user.status}')

        # Синхронная отправка email
        send_mail(
            'Приглашение на платформу Task Manager',
            f'Ваши учетные данные для входа:\nЛогин: {user.username}\nПароль: {password}',
            'info@demotimetracker.ru',
            [user.email],
            fail_silently=False,
        )

        print('Приглашение отправлено успешно.')

        messages.success(request, 'Приглашение отправлено успешно.')
    else:
        messages.warning(request, 'Этот пользователь уже был приглашен ранее.')

    return redirect('user-list')


def reset_password(request, pk):
    user = get_object_or_404(CustomUser, pk=pk)

    # Генерация нового случайного пароля
    new_password = get_random_string(length=8)
    user.set_password(new_password)
    user.save()

    # Отправка нового пароля на email пользователя
    send_mail(
        'Ваш новый пароль',
        f'Ваш новый пароль: {new_password}',
        'noreply@taskmanager.com',  # Укажите email отправителя
        [user.email],
        fail_silently=False,
    )

    messages.success(request, 'Пароль успешно сброшен и отправлен на email пользователя.')
    return redirect('user-list')


# Проекты
@admin_required(login_url='login')
def project(request):
    # Кэшируем проекты и связанные данные
    projects = cache.get_or_set(
        'projects_cache',
        Project.objects.select_related('status').prefetch_related('project_buildings__building',
                                                                  'project_sections__section').annotate(
            timelog_count=Count('timelogs')),
        timeout=60 * 15  # Кэш на 15 минут
    )

    project_status = cache.get_or_set(
        'project_status_cache',
        ProjectStatus.objects.all(),
        timeout=60 * 15
    )

    sections = cache.get_or_set(
        'sections_cache',
        Section.objects.all(),
        timeout=60 * 15
    )

    # Фильтры из GET-запроса
    selected_projects = request.GET.getlist('project')
    selected_sections = request.GET.getlist('section')
    selected_statuses = request.GET.getlist('status')
    selected_buildings = request.GET.getlist('zis')

    # Создаем Q-объекты для фильтров
    filter_conditions = Q()

    if selected_projects:
        selected_projects = selected_projects[0].split(',')
        filter_conditions &= Q(title__in=selected_projects)
    if selected_sections:
        selected_sections = selected_sections[0].split(',')
        filter_conditions &= Q(project_sections__section__title__in=selected_sections)
    if selected_statuses:
        selected_statuses = selected_statuses[0].split(',')
        status_ids = ProjectStatus.objects.filter(title__in=selected_statuses).values_list('id', flat=True)
        filter_conditions &= Q(status_id__in=status_ids)
    if selected_buildings:
        selected_buildings = selected_buildings[0].split(',')
        filter_conditions &= Q(project_buildings__building__title__in=selected_buildings)

    # Применяем фильтры
    projects = projects.filter(filter_conditions)

    form = ProjectForm()

    return render(request, 'task_manager/project_list.html', {
        'projects': projects,
        'project_status': project_status,
        'sections': sections,
        'form': form
    })


def project_detail(request, pk):
    try:
        cache_key = f'project_detail_{pk}'
        project_data = cache.get(cache_key)

        if not project_data:
            project = Project.objects.prefetch_related(
                'project_buildings__building',
                'project_sections__section'
            ).get(pk=pk)

            buildings = [
                {
                    'building__pk': building['building__pk'],
                    'building__title': building['building__title'].strip()  # Remove extra whitespace or newlines
                }
                for building in project.project_buildings.values('building__pk', 'building__title')
            ]

            sections = [
                {
                    'section__pk': section['section__pk'],
                    'section__title': section['section__title']
                }
                for section in project.project_sections.values('section__pk', 'section__title')
            ]

            project_data = {
                'project': {
                    'title': project.title,
                    'status': str(project.status_id)  # Ensure UUID is serialized as a string
                },
                'buildings': buildings,
                'sections': sections
            }
            cache.set(cache_key, project_data, timeout=60 * 15)

        return JsonResponse(project_data)

    except Project.DoesNotExist:
        return JsonResponse({'error': 'Project not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def createProject(request):
    if request.method == 'POST':
        title = request.POST['title']
        status_id = request.POST['status']
        sections = request.POST.getlist('sections[]')

        with transaction.atomic():
            project = Project.objects.create(title=title, status_id=status_id)

            # Обработка зданий
            buildings = request.POST.getlist('buildings[]')
            for building_title in buildings:
                building, created = Building.objects.get_or_create(title=building_title)
                ProjectBuilding.objects.create(project=project, building=building)

            # Добавляем выбранные разделы к проекту
            for section_id in sections:
                section = Section.objects.get(pk=section_id)
                ProjectSection.objects.create(project=project, section=section)

            # Очищаем кэш после добавления проекта
            cache.delete('projects_cache')
            cache.delete('project_status_cache')
            cache.delete('sections_cache')

        return JsonResponse({'message': 'Проект создан'})


def updateProject(request, pk):
    if request.method == 'POST':
        project = get_object_or_404(Project, pk=pk)

        with transaction.atomic():
            project.title = request.POST['title']
            project.status_id = request.POST['status']
            project.save()

            buildings_ids = request.POST.getlist('buildings_id[]')
            buildings_titles = request.POST.getlist('buildings[]')
            print(f"buildings_ids: {buildings_ids}")
            print(f"buildings_titles: {buildings_titles}")

            n_ids = len(buildings_ids)
            for i, building_title in enumerate(buildings_titles):
                if i < n_ids and buildings_ids[i]:
                    building_id = buildings_ids[i]
                    if building_id:
                        try:
                            building = Building.objects.get(pk=building_id)
                            print(f"Обновляем здание {building_id} с новым названием '{building_title}'")
                            building.title = building_title
                            building.save()
                        except Building.DoesNotExist:
                            print(f"Здание с id {building_id} не найдено, создаём новое с названием '{building_title}'")
                            building, created = Building.objects.get_or_create(title=building_title)
                    else:
                        print(
                            f"Передан некорректный UUID {building_id}, создаём новое здание с названием '{building_title}'")
                        building, created = Building.objects.get_or_create(title=building_title)
                else:
                    print(f"Нет id для здания с названием '{building_title}', создаём новое")
                    building, created = Building.objects.get_or_create(title=building_title)

                ProjectBuilding.objects.update_or_create(
                    project=project,
                    building=building,
                    defaults={}
                )

            # Если требуется удалять связи с удаленными зданиями:
            removed_buildings = request.POST.getlist('removed_buildings[]')
            for building_id in removed_buildings:
                try:
                    project_building = ProjectBuilding.objects.get(project=project, building_id=building_id)
                    project_building.delete()

                    # Удаляем само здание, если оно не связано с другими проектами
                    building = Building.objects.get(pk=building_id)
                    if not ProjectBuilding.objects.filter(building=building).exists():
                        building.delete()
                except ProjectBuilding.DoesNotExist:
                    continue

            # Обработка разделов
            sections = request.POST.getlist('sections[]')
            ProjectSection.objects.filter(project=project).delete()
            for section_id in sections:
                section = Section.objects.get(pk=section_id)
                ProjectSection.objects.create(project=project, section=section)

            # Очищаем кэш после обновления проекта
            cache.delete('projects_cache')
            cache.delete(f'project_detail_{pk}')
            cache.delete('project_status_cache')
            cache.delete('sections_cache')

        return JsonResponse({'message': 'Проект обновлен'})


def deleteProject(request, pk):
    if request.method == 'POST':
        project = get_object_or_404(Project, pk=pk)
        project.delete()

        # Очищаем кэш после удаления проекта
        cache.delete('projects_cache')
        cache.delete(f'project_detail_{pk}')
        cache.delete('project_status_cache')
        cache.delete('sections_cache')

        return JsonResponse({'message': 'Проект удален'})


# ProjectBuilding

def project_building_create(request, project_pk):
    project = get_object_or_404(Project, pk=project_pk)
    if request.method == 'POST':
        form = ProjectBuildingForm(request.POST)
        if form.is_valid():
            project_building = form.save(commit=False)
            project_building.project = project
            project_building.save()
            return redirect('project-detail', pk=project_pk)
    else:
        form = ProjectBuildingForm()
    return render(request, 'task_manager/project_building_form.html', {'form': form, 'project': project})


def project_building_delete(request, pk):
    project_building = get_object_or_404(ProjectBuilding, pk=pk)
    project_pk = project_building.project.pk
    if request.method == 'POST':
        project_building.delete()
        return redirect('project-detail', pk=project_pk)
    return render(request, 'task_manager/project_building_confirm_delete.html', {'project_building': project_building})


# ProjectSection

def project_section_create(request, project_pk):
    project = get_object_or_404(Project, pk=project_pk)
    if request.method == 'POST':
        form = ProjectSectionForm(request.POST)
        if form.is_valid():
            project_section = form.save(commit=False)
            project_section.project = project
            project_section.save()
            return redirect('project-detail', pk=project_pk)
    else:
        form = ProjectSectionForm()
    return render(request, 'task_manager/project_section_form.html', {'form': form, 'project': project})


def project_section_delete(request, pk):
    project_section = get_object_or_404(ProjectSection, pk=pk)
    project_pk = project_section.project.pk
    if request.method == 'POST':
        project_section.delete()
        return redirect('project-detail', pk=project_pk)
    return render(request, 'task_manager/project_section_confirm_delete.html', {'project_section': project_section})


def building_create(request):
    if request.method == 'POST':
        building_title = request.POST.get('title')
        project_id = request.POST.get('project_id')

        try:
            project = Project.objects.get(pk=project_id)
        except Project.DoesNotExist:
            return JsonResponse({'error': 'Project not found'}, status=404)

        try:
            building, created = Building.objects.get_or_create(title=building_title)

            if ProjectBuilding.objects.filter(project=project, building=building).exists():
                return JsonResponse({'error': 'Building already exists in this project'}, status=400)

            ProjectBuilding.objects.create(project=project, building=building)
        except IntegrityError:
            return JsonResponse({'error': 'Building with this name already exists'}, status=400)

        return JsonResponse({'message': 'Building created and added to project'}, status=200)

    return JsonResponse({'error': 'Invalid request'}, status=400)


def building_delete(request, building_id):
    if request.method == 'POST':
        project_id = request.POST.get('project_id')

        # Удаление связи между проектом и зданием
        try:
            project_building = get_object_or_404(ProjectBuilding, building_id=building_id, project_id=project_id)
            project_building.delete()

            # Теперь удаляем само здание, если оно не связано с другими проектами
            building = get_object_or_404(Building, pk=building_id)
            # Проверяем, связано ли здание с другими проектами
            if not ProjectBuilding.objects.filter(building=building).exists():
                building.delete()

            return JsonResponse({'message': 'Building removed from project and deleted'})
        except Exception as e:
            return JsonResponse({'error': 'Error deleting building: ' + str(e)}, status=400)


# Разделы
@admin_required(login_url='login')
def section(request):
    sections = Section.objects.annotate(timelog_count=Count('timelogs', distinct=True),
                                        project_count=Count('project_sections', distinct=True))
    available_marks = Mark.objects.all()  # Получаем все доступные марки

    selected_sections = request.GET.getlist('section')

    if selected_sections:
        selected_sections = selected_sections[0].split(',')
        sections = sections.filter(title__in=selected_sections)

    form = SectionForm()

    return render(request, 'task_manager/sections_list.html', {
        'sections': sections,
        'form': form,
        'available_marks': available_marks  # Передаем доступные марки в контекст
    })


def createSection(request):
    form = SectionForm()

    if request.method == 'POST':
        form = SectionForm(request.POST)
        if form.is_valid():
            form.save()
            cache.delete('sections_cache')
            return redirect('section-list')

    return render(request, 'task_manager/section_form.html', {'form': form})


def updateSection(request, pk):
    section = get_object_or_404(Section, pk=pk)
    available_marks = Mark.objects.all()  # Все доступные марки
    section_marks = SectionMark.objects.filter(section=section)  # Марки, связанные с секцией

    if request.method == 'POST':
        form = SectionForm(request.POST, instance=section)
        if form.is_valid():
            form.save()
            cache.delete('sections_cache')

            # Обновляем марки для секции
            marks = request.POST.getlist('marks')  # Получаем выбранные ID марок из формы
            SectionMark.objects.filter(section=section).delete()  # Удаляем старые марки

            for mark_id in marks:
                mark = Mark.objects.get(id=mark_id)
                SectionMark.objects.create(section=section, mark=mark)  # Создаем новые связи

            return redirect('section-list')
    else:
        form = SectionForm(instance=section)

    return render(request, 'task_manager/section_form.html', {
        'form': form,
        'section': section,
        'available_marks': available_marks,
        'section_marks': section_marks,
    })


def deleteSection(request, pk):
    section = get_object_or_404(Section, pk=pk)
    if request.method == 'POST':
        if section.project_sections.exists() or section.timelogs.exists():
            return redirect('section-list')

        section.delete()
        cache.delete('sections_cache')
        return redirect('section-list')

    return render(request, 'task_manager/section_confirm_delete.html', {'section': section})


# Марки
@admin_required(login_url='login')
def mark(request):
    # Попробуем получить марки из кэша
    marks = cache.get('mark_list')
    if not marks:
        # Если кэш пуст, загружаем данные из базы и кэшируем
        marks = Mark.objects.annotate(
            department_id=Subquery(
                DepartmentMark.objects.filter(mark=OuterRef('pk')).values('department')[:1]
            )
        ).annotate(
            departments=ArrayAgg(
                'department_marks__department__title',
                distinct=True,
                filter=Q(department_marks__department__title__isnull=False)
            ),
            timelog_count=Count('timelogs', distinct=True),
            section_count=Count('section_marks', distinct=True)
        )
        cache.set('mark_list', marks, timeout=60 * 15)  # Кэшируем на 15 минут

    selected_marks = request.GET.getlist('mark')

    if selected_marks:
        selected_marks = selected_marks[0].split(',')
        marks = marks.filter(title__in=selected_marks)

    # Фильтрация по названию отдела
    selected_departments = request.GET.getlist('department')
    # print(selected_departments)
    if selected_departments:
        if len(selected_departments) == 1 and ',' in selected_departments[0]:
            selected_departments = selected_departments[0].split(',')
        # Фильтрация через связь DepartmentMark -> Department -> title
        marks = marks.filter(department_marks__department__title__in=selected_departments)

    # Из-за соединения по связанной таблице могут возникать дубликаты – оставляем только уникальные объекты
    # marks = marks.distinct()

    form = MarkForm()
    return render(request, 'task_manager/mark_list.html', {'marks': marks, 'form': form})


def createMark(request):
    form = MarkForm()
    if request.method == 'POST':
        form = MarkForm(request.POST)
        if form.is_valid():
            mark = form.save(commit=True)  # Сохраняем саму Mark

            # Получаем отдел, который выбрал пользователь
            chosen_department = form.cleaned_data['department']

            # Удаляем старые связи (по идее для create их и нет, но на всякий случай)
            DepartmentMark.objects.filter(mark=mark).delete()

            # Создаём новую связь, если пользователь выбрал отдел
            if chosen_department:
                DepartmentMark.objects.create(mark=mark, department=chosen_department)

            cache.delete('mark_list')  # Очистим кэш
            return redirect('mark-list')

    context = {'form': form}
    return render(request, 'task_manager/mark_form.html', context)


def updateMark(request, pk):
    mark = get_object_or_404(Mark, pk=pk)
    if request.method == 'POST':
        form = MarkForm(request.POST, instance=mark)
        if form.is_valid():
            mark = form.save(commit=True)

            chosen_department = form.cleaned_data['department']
            # Удаляем старую связь (если была)
            DepartmentMark.objects.filter(mark=mark).delete()

            if chosen_department:
                DepartmentMark.objects.create(mark=mark, department=chosen_department)
            else:
                DepartmentMark.objects.filter(mark=mark).delete()

            cache.delete('mark_list')
            return redirect('mark-list')

    else:
        form = MarkForm(instance=mark)

    context = {'form': form, 'mark': mark}
    return render(request, 'task_manager/mark_form.html', context)


def deleteMark(request, pk):
    mark = get_object_or_404(Mark, pk=pk)
    if request.method == 'POST':
        if mark.section_marks.exists() or mark.timelogs.exists():
            return redirect('mark-list')

        # Удаляем связку с отделом, если она существует
        DepartmentMark.objects.filter(mark=mark).delete()

        mark.delete()
        cache.delete('mark_list')  # Удаляем кэш после удаления
        return redirect('mark-list')

    context = {'object': mark}
    return render(request, 'task_manager/mark_confirm_delete.html', context)


# Задачи
@admin_required(login_url='login')
def task_list(request):
    # Попробуем получить задачи из кэша
    cache_key = 'tasktype_list'
    tasks = cache.get(cache_key)

    if not tasks:
        tasks = TaskType.objects.annotate(
            department_id=Subquery(
                DepartmentTaskType.objects.filter(task=OuterRef('pk')).values('department')[:1]
            )
        ).annotate(
            timelog_count=Count('timelogs', distinct=True),
            departments=ArrayAgg(
                'department_tasks__department__title',
                distinct=True,
                filter=Q(department_tasks__department__title__isnull=False)
            ),
        )

    selected_tasks = request.GET.getlist('tasks')

    if selected_tasks:
        selected_tasks = selected_tasks[0].split(',')
        tasks = tasks.filter(title__in=selected_tasks)

    # Фильтрация по названию отдела
    selected_departments = request.GET.getlist('department')
    # print(selected_departments)
    if selected_departments:
        if len(selected_departments) == 1 and ',' in selected_departments[0]:
            selected_departments = selected_departments[0].split(',')
        # Фильтрация через связь DepartmentTaskType -> Department -> title
        tasks = tasks.filter(department_tasks__department__title__in=selected_departments)

    cache.set(cache_key, tasks, timeout=60 * 15)  # Кэшируем на 15 минут

    form = TaskTypeForm()

    return render(request, 'task_manager/task_list.html', {'tasks': tasks, 'form': form})


def create_task(request):
    form = TaskTypeForm()

    if request.method == 'POST':
        form = TaskTypeForm(request.POST)
        if form.is_valid():
            task = form.save(commit=False)
            task.save()

            # Если есть привязка отдела, создаем запись в DepartmentTaskType
            department_id = form.cleaned_data['department']
            if department_id:
                DepartmentTaskType.objects.create(department=department_id, task=task)

            cache.delete('tasktype_list')  # Удаляем кэш после создания
            return redirect('task-list')

    return render(request, 'task_manager/task_form.html', {'form': form})


def update_task(request, pk):
    task = get_object_or_404(TaskType, pk=pk)
    form = TaskTypeForm(instance=task)

    if request.method == 'POST':
        form = TaskTypeForm(request.POST, instance=task)
        if form.is_valid():
            task = form.save()
            department_id = form.cleaned_data['department']

            # Удаляем все старые записи для этой задачи
            DepartmentTaskType.objects.filter(task=task).delete()

            # Добавляем новую связь, если отдел указан
            if department_id:
                DepartmentTaskType.objects.create(task=task, department=department_id)

            cache.delete('tasktype_list')  # Удаляем кэш после обновления
            return redirect('task-list')

    return render(request, 'task_manager/task_form.html', {'form': form, 'task': task})


def delete_task(request, pk):
    task = get_object_or_404(TaskType, pk=pk)

    if request.method == 'POST':
        if task.timelogs.exists():
            return redirect('task-list')

        # Удаляем связку с отделом, если она существует
        DepartmentTaskType.objects.filter(task=task).delete()

        task.delete()
        cache.delete('tasktype_list')  # Удаляем кэш после удаления
        return redirect('task-list')

    return render(request, 'task_manager/task_confirm_delete.html', {'task': task})


def parse_custom_date(date_str):
    # Убираем день недели и удаляем символ "г." в конце, если он есть
    date_str = date_str.split(",")[-1].strip()
    date_str = date_str.replace(" г.", "").strip()

    # Попробуем сначала стандартный формат "YYYY-MM-DD"
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        pass

    # Попробуем формат с русскими названиями месяцев
    months = {
        "января": "01", "февраля": "02", "марта": "03", "апреля": "04",
        "мая": "05", "июня": "06", "июля": "07", "августа": "08",
        "сентября": "09", "октября": "10", "ноября": "11", "декабря": "12"
    }

    # Обработка формата "дд ммм гггг"
    match = re.match(r"(\d{1,2}) (\w+) (\d{4})", date_str)
    if match:
        day, month_str, year = match.groups()
        month = months.get(month_str.lower())
        if month:
            return datetime.strptime(f"{year}-{month}-{day.zfill(2)}", "%Y-%m-%d").date()

    # Обработка формата "дд.мм.гггг"
    try:
        return datetime.strptime(date_str, "%d.%m.%Y").date()
    except ValueError:
        pass

    # Если не удалось распознать формат
    raise ValueError(f"Unexpected date format: {date_str}")


# TimeLog
@admin_required(login_url='login')
def timelog_list(request):
    # Получаем начальную и конечную даты из GET параметров
    start_date = request.GET.get('start_date', timezone.now().replace(day=1).date())
    end_date = request.GET.get('end_date', timezone.now().date())

    # Преобразуем даты с использованием parse_custom_date, если они строковые
    start_date = parse_custom_date(start_date) if isinstance(start_date, str) else start_date
    end_date = parse_custom_date(end_date) if isinstance(end_date, str) else end_date

    # Фильтры из GET-запроса
    selected_projects = request.GET.getlist('project')
    selected_departments = request.GET.getlist('department')
    selected_users = request.GET.getlist('employees')
    selected_stages = request.GET.getlist('stage')
    selected_sections = request.GET.getlist('section')
    selected_buildings = request.GET.getlist('zis')
    selected_marks = request.GET.getlist('mark')
    selected_tasks = request.GET.getlist('task')

    # Кэшируем данные таймлогов, если кэш пуст
    timelogs = cache.get('timelogs_cache')
    if not timelogs:
        timelogs = (
            Timelog.objects.all()
            .select_related('user', 'role', 'department', 'project', 'building', 'mark', 'task')
            .prefetch_related('section')
        )
        # Сохраняем в кэше на 5 минут (300 секунд)
        cache.set('timelogs_cache', timelogs, timeout=900)

    # Применяем фильтры, если они указаны
    if selected_projects:
        selected_projects = selected_projects[0].split(',')
        timelogs = timelogs.filter(project__title__in=selected_projects)
    if selected_sections:
        selected_sections = selected_sections[0].split(',')
        timelogs = timelogs.filter(section__title__in=selected_sections)
    if selected_departments:
        selected_departments = selected_departments[0].split(',')
        timelogs = timelogs.filter(department__title__in=selected_departments)
    if selected_users:
        user_filters = Q()
        user_list = selected_users[0].split(',')
        for full_name in user_list:
            first_name, last_name = full_name.strip().split(' ', 1)
            user_filters |= Q(user__first_name=first_name, user__last_name=last_name)
        timelogs = timelogs.filter(user_filters)
    if selected_stages:
        selected_stages = selected_stages[0].split(',')
        timelogs = timelogs.filter(stage__in=selected_stages)
    if selected_buildings:
        selected_buildings = selected_buildings[0].split(',')
        timelogs = timelogs.filter(building__title__in=selected_buildings)
    if selected_marks:
        selected_marks = selected_marks[0].split(',')
        timelogs = timelogs.filter(mark__title__in=selected_marks)
    if selected_tasks:
        selected_tasks = selected_tasks[0].split(',')
        timelogs = timelogs.filter(task__title__in=selected_tasks)

    # Фильтрация по диапазону дат
    if start_date and end_date:
        timelogs = timelogs.filter(date__range=[start_date, end_date])

    form = TimelogForm()

    context = {
        'timelogs': timelogs,
        'form': form,
        'start_date': start_date,
        'end_date': end_date,
    }

    return render(request, 'task_manager/timelog_list.html', context)


@admin_required(login_url='login')
def timelog_create(request):
    if request.method == 'POST':
        form = TimelogForm(request.POST)
        if form.is_valid():
            # Получаем текущего пользователя
            user = request.user
            role = user.role  # Предполагается, что у пользователя есть роль

            # Создаем новый таймлог с указанием пользователя и его роли
            timelog = form.save(commit=False)
            timelog.user = user
            timelog.role = role
            timelog.save()
            cache.delete('timelogs_cache')

            return JsonResponse({'message': 'Таймлог успешно создан'}, status=200)

    return JsonResponse({'error': 'Неверный запрос'}, status=400)


def timelog_update(request, pk):
    timelog = get_object_or_404(Timelog, pk=pk)
    form = TimelogForm(instance=timelog)

    if request.method == 'POST':
        form = TimelogForm(request.POST, instance=timelog)
        if form.is_valid():
            form.save()
            cache.delete('timelogs_cache')
            return redirect('timelog-list')

    context = {'form': form}
    return render(request, 'task_manager/timelog_form.html', context)


def timelog_delete(request, pk):
    timelog = get_object_or_404(Timelog, pk=pk)
    if request.method == 'POST':
        timelog.delete()
        return redirect('timelog-list')

    context = {'object': timelog}
    return render(request, 'task_manager/timelog_confirm_delete.html', context)


def reset_session(request):
    request.session.flush()  # Полностью очистит текущую сессию
    return redirect('login')


@login_required
def user_dashboard(request):
    user = request.user
    moscow_tz = pytz.timezone("Europe/Moscow")

    # Получаем диапазон всех дат от момента создания пользователя до текущей даты
    start_date = user.created_at.astimezone(moscow_tz).date()
    end_date = now().astimezone(moscow_tz).date()
    dates_range = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]

    # Получаем уже заполненные таймлоги пользователя и суммируем время по датам
    filled_timelogs = Timelog.objects.filter(user=user) \
        .annotate(date_only=TruncDate('date')) \
        .values('date_only') \
        .annotate(total_time=Sum('time')) \
        .values('date_only', 'total_time')

    # Преобразуем в словарь: {дата: суммарное время}
    filled_dates = {log['date_only']: log['total_time'] for log in filled_timelogs}

    # Формируем список отчетов
    reports_to_fill = [
        {"date": date, "is_filled": date in filled_dates, "total_time": filled_dates.get(date, 0)}
        for date in dates_range
    ]

    # Фильтрация отчетов за последнюю неделю, которые были заполнены
    one_week_ago = timezone.now().date() - timedelta(days=7)
    filled_last_week = [
        {"date": date, "is_filled": date in filled_dates, "total_time": filled_dates.get(date, 0)}
        for date in dates_range
        if date >= one_week_ago and date in filled_dates
    ]

    # Получаем только незаполненные отчеты и отчеты, заполненные за последнюю неделю
    reports_to_display = [
                             report for report in reports_to_fill if not report["is_filled"]
                         ] + filled_last_week

    if request.method == 'POST':
        # Обработка отправки таймлога
        date_to_fill = request.POST.get('date')
        if date_to_fill:
            form = TimelogForm(request.POST)
            if form.is_valid():
                new_timelog = form.save(commit=False)
                new_timelog.user = user
                new_timelog.role = user.role
                new_timelog.date = date_to_fill
                new_timelog.save()
                return redirect('user-dashboard')

    context = {
        'user': user,
        'reports_to_display': reports_to_display,
    }
    return render(request, 'task_manager/user_dashboard.html', context)


def get_buildings(request):
    project_id = request.GET.get('project_id')  # Получаем project_id из запроса
    if not project_id:
        return JsonResponse({'error': 'Missing project_id'}, status=400)

    try:
        # Используем ProjectBuilding для связи между проектом и зданиями
        buildings = ProjectBuilding.objects.filter(project_id=project_id).select_related('building')
        buildings_data = [{'id': pb.building.id, 'title': pb.building.title} for pb in buildings]
        return JsonResponse({'buildings': buildings_data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def get_sections(request):
    project_id = request.GET.get('project_id')  # Получаем project_id из запроса
    if not project_id:
        return JsonResponse({'error': 'Missing project_id'}, status=400)

    try:
        # Используем ProjectSection для связи между проектом и разделами
        sections = ProjectSection.objects.filter(project_id=project_id).select_related('section')
        sections_data = [{'id': ps.section.id, 'title': ps.section.title} for ps in sections]
        return JsonResponse({'sections': sections_data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def natural_keys(text):
    """
    Разбивает строку на список числовых и нечисловых частей,
    чтобы сортировка шла в «естественном» порядке.
    Пример: "Task 2" будет меньше чем "Task 10".
    """
    return [int(c) if c.isdigit() else c.lower() for c in re.split(r'(\d+)', text)]


@login_required
def report_create(request):
    # Получаем дату из запроса или используем текущую
    requested_date = request.GET.get('date')

    if requested_date:
        selected_date = parse_russian_date(requested_date)  # Используем ваш парсер
        if not selected_date:
            selected_date = timezone.now().date()  # Если не удалось распарсить, используем текущую
    else:
        selected_date = timezone.now().date()

    # Если отчет редактируется, получаем уже существующие записи
    timelogs_data = []
    if request.method == 'GET' and requested_date:
        # Попробуем найти существующие таймлоги для данной даты
        timelogs_data = Timelog.objects.filter(user=request.user, date=selected_date)

    if request.method == 'POST':
        projects = request.POST.getlist('project')
        stages = request.POST.getlist('stage')
        sections = request.POST.getlist('section')
        buildings = request.POST.getlist('building')
        marks = request.POST.getlist('mark')
        tasks = request.POST.getlist('task')
        times = request.POST.getlist('time')
        post_date = request.POST.get('date')

        # 1. Удаляем все старые таймлоги для данной даты
        Timelog.objects.filter(user=request.user, date=post_date).delete()

        # 2. Собираем данные для новых таймлогов
        timelogs_data = []
        for i in range(len(projects)):
            task_title = tasks[i].strip()
            task, _ = TaskType.objects.get_or_create(title=task_title)

            timelog = Timelog(
                user=request.user,
                role=request.user.role,
                department=request.user.department,
                project_id=projects[i],
                stage=stages[i],
                section_id=sections[i],
                building_id=buildings[i],
                mark_id=marks[i],
                task_id=task.id,
                time=int(times[i]),
                date=post_date
            )
            timelogs_data.append(timelog)

        # 3. Сохраняем все новые таймлоги одним запросом
        Timelog.objects.bulk_create(timelogs_data)

        # Очищаем кэш, чтобы подтянуть актуальные данные при следующем запросе
        cache.delete('projects_cache')
        cache.delete('sections_cache')
        cache.delete('buildings')
        cache.delete('mark_list')
        cache.delete('tasks')
        cache.delete('timelogs_cache')

        return redirect('user-dashboard')  # Перенаправляем на список таймлогов

    else:
        form = TimelogForm()

    # Загружаем данные из кэша или базы данных для селектов
    projects = cache.get_or_set('projects_cache', Project.objects.select_related('status').all(), timeout=60 * 15)
    sections = cache.get_or_set('sections_cache', Section.objects.all(), timeout=60 * 15)
    marks = Mark.objects.filter(department_marks__department=request.user.department)
    tasks = TaskType.objects.filter(department_tasks__department=request.user.department)

    # Применяем натуральную сортировку по названию
    projects = sorted(
        cache.get_or_set('projects_cache', Project.objects.select_related('status').all(), timeout=60 * 15),
        key=lambda project: natural_keys(project.title)
    )

    sections = sorted(
        cache.get_or_set('sections_cache', Section.objects.all(), timeout=60 * 15),
        key=lambda section: natural_keys(section.title)
    )

    marks = sorted(
        Mark.objects.filter(department_marks__department=request.user.department),
        key=lambda mark: natural_keys(mark.title)
    )

    tasks = sorted(
        TaskType.objects.filter(department_tasks__department=request.user.department),
        key=lambda task: natural_keys(task.title)
    )

    context = {
        'form': form,
        'selected_date': selected_date,
        'projects': projects,
        'stages': Timelog._meta.get_field('stage').choices,
        'sections': sections,
        'buildings': Building.objects.all(),
        'marks': marks,
        'tasks': tasks,
        'timelogs_data': timelogs_data,
    }

    return render(request, 'task_manager/report_create.html', context)


@admin_required(login_url='login')
def reports_view(request):
    # Получаем начальную и конечную даты из GET параметров
    start_date = request.GET.get('start_date', timezone.now().replace(day=1).date())
    end_date = request.GET.get('end_date', timezone.now().date())

    # Преобразуем даты с использованием новой функции
    start_date = parse_custom_date(start_date) if isinstance(start_date, str) else start_date
    end_date = parse_custom_date(end_date) if isinstance(end_date, str) else end_date

    # Фильтры из GET-запроса
    selected_projects = request.GET.getlist('project')
    selected_departments = request.GET.getlist('department')
    selected_users = request.GET.getlist('employees')
    selected_stages = request.GET.getlist('stage')
    selected_buildings = request.GET.getlist('zis')
    selected_marks = request.GET.getlist('mark')
    selected_tasks = request.GET.getlist('task')

    # Фильтрация Timelog по диапазону дат
    timelogs = (
        Timelog.objects
        .filter(date__range=[start_date, end_date])
        .select_related('project', 'department', 'user', 'building', 'mark', 'task')
    )

    # Применяем фильтры, если они указаны
    if selected_projects:
        selected_projects = selected_projects[0].split(',')
        timelogs = timelogs.filter(project__title__in=selected_projects)
    if selected_departments:
        selected_departments = selected_departments[0].split(',')
        timelogs = timelogs.filter(department__title__in=selected_departments)
    if selected_users:
        user_filters = Q()
        user_list = selected_users[0].split(',')
        for full_name in user_list:
            first_name, last_name = full_name.strip().split(' ', 1)
            user_filters |= Q(user__first_name=first_name, user__last_name=last_name)
        timelogs = timelogs.filter(user_filters)
    if selected_stages:
        selected_stages = selected_stages[0].split(',')
        timelogs = timelogs.filter(stage__in=selected_stages)
    if selected_buildings:
        selected_buildings = selected_buildings[0].split(',')
        timelogs = timelogs.filter(building__title__in=selected_buildings)
    if selected_marks:
        selected_marks = selected_marks[0].split(',')
        timelogs = timelogs.filter(mark__title__in=selected_marks)
    if selected_tasks:
        selected_tasks = selected_tasks[0].split(',')
        timelogs = timelogs.filter(task__title__in=selected_tasks)

    # Группировка данных по проектам
    detailed_report_projects = {}
    for item in timelogs:
        project_title = item.project.title
        if project_title not in detailed_report_projects:
            detailed_report_projects[project_title] = {
                'entries': [],
                'total_time': 0
            }
        detailed_report_projects[project_title]['entries'].append(item)
        detailed_report_projects[project_title]['total_time'] += item.time

    # Общий итог времени по всем проектам
    overall_total_time_projects = sum([group['total_time'] for group in detailed_report_projects.values()])

    # Группировка данных по отделам
    detailed_report_departments = {}
    for item in timelogs:
        department_title = item.department.title
        if department_title not in detailed_report_departments:
            detailed_report_departments[department_title] = {
                'entries': [],
                'total_time': 0
            }
        detailed_report_departments[department_title]['entries'].append(item)
        detailed_report_departments[department_title]['total_time'] += item.time

    # Общий итог времени по всем отделам
    overall_total_time_departments = sum([group['total_time'] for group in detailed_report_departments.values()])

    context = {
        'detailed_report_projects': detailed_report_projects,
        'overall_total_time_projects': overall_total_time_projects,
        'detailed_report_departments': detailed_report_departments,
        'overall_total_time_departments': overall_total_time_departments,
        'start_date': start_date,
        'end_date': end_date
    }
    return render(request, 'task_manager/reports.html', context)


@admin_required(login_url='login')
def reports_employees(request):
    # Получаем начальную и конечную даты из GET параметров
    start_date = request.GET.get('start_date', timezone.now().replace(day=1).date())
    end_date = request.GET.get('end_date', timezone.now().date())

    # Преобразуем даты с использованием новой функции
    start_date = parse_custom_date(start_date) if isinstance(start_date, str) else start_date
    end_date = parse_custom_date(end_date) if isinstance(end_date, str) else end_date

    # Фильтры из GET-запроса
    selected_projects = request.GET.getlist('project')
    selected_departments = request.GET.getlist('department')
    selected_users = request.GET.getlist('employees')
    selected_stages = request.GET.getlist('stage')
    selected_buildings = request.GET.getlist('zis')
    selected_marks = request.GET.getlist('mark')
    selected_tasks = request.GET.getlist('task')

    # Получаем все записи Timelog с оптимизацией связанных данных
    timelogs = (
        Timelog.objects
        .filter(date__range=[start_date, end_date])
        .select_related('project', 'department', 'user', 'building', 'mark', 'task')
    )

    # Применяем фильтры, если они указаны
    if selected_projects:
        selected_projects = selected_projects[0].split(',')
        timelogs = timelogs.filter(project__title__in=selected_projects)
    if selected_departments:
        selected_departments = selected_departments[0].split(',')
        timelogs = timelogs.filter(department__title__in=selected_departments)
    if selected_users:
        user_filters = Q()
        user_list = selected_users[0].split(',')
        for full_name in user_list:
            first_name, last_name = full_name.strip().split(' ', 1)
            user_filters |= Q(user__first_name=first_name, user__last_name=last_name)
        timelogs = timelogs.filter(user_filters)
    if selected_stages:
        selected_stages = selected_stages[0].split(',')
        timelogs = timelogs.filter(stage__in=selected_stages)
    if selected_buildings:
        selected_buildings = selected_buildings[0].split(',')
        timelogs = timelogs.filter(building__title__in=selected_buildings)
    if selected_marks:
        selected_marks = selected_marks[0].split(',')
        timelogs = timelogs.filter(mark__title__in=selected_marks)
    if selected_tasks:
        selected_tasks = selected_tasks[0].split(',')
        timelogs = timelogs.filter(task__title__in=selected_tasks)

    # Группировка данных по проектам
    detailed_report_projects = {}
    for item in timelogs:
        project_title = item.project.title
        if project_title not in detailed_report_projects:
            detailed_report_projects[project_title] = {
                'entries': [],
                'total_time': 0
            }
        detailed_report_projects[project_title]['entries'].append(item)
        detailed_report_projects[project_title]['total_time'] += item.time

    # Общий итог времени по всем проектам
    overall_total_time_projects = sum(group['total_time'] for group in detailed_report_projects.values())

    # Группировка данных по сотрудникам
    detailed_report_employees = {}
    for item in timelogs:
        employee_full_name = f"{item.user.first_name} {item.user.last_name}"
        if employee_full_name not in detailed_report_employees:
            detailed_report_employees[employee_full_name] = {
                'entries': [],
                'total_time': 0
            }
        detailed_report_employees[employee_full_name]['entries'].append(item)
        detailed_report_employees[employee_full_name]['total_time'] += item.time

    # Общий итог времени по всем сотрудникам
    overall_total_time_employees = sum(group['total_time'] for group in detailed_report_employees.values())

    context = {
        'detailed_report_projects': detailed_report_projects,
        'overall_total_time_projects': overall_total_time_projects,
        'detailed_report_departments': detailed_report_employees,
        'overall_total_time_departments': overall_total_time_employees,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'task_manager/reports_employees.html', context)


def get_months_in_range(start_date, end_date):
    """Получение всех месяцев до последнего месяца в диапазоне"""
    months = []
    current_date = start_date.replace(day=1)
    while current_date <= end_date.replace(day=1):
        months.append(current_date)
        if current_date.month == 12:
            current_date = current_date.replace(year=current_date.year + 1, month=1)
        else:
            current_date = current_date.replace(month=current_date.month + 1)
    return months[:-1]  # Все месяцы, кроме последнего


def get_days_in_month(month_date):
    """Получение всех дней для последнего месяца"""
    next_month = month_date.replace(day=28) + timedelta(days=4)
    end_of_month = next_month - timedelta(days=next_month.day)
    return [month_date + timedelta(days=i) for i in range((end_of_month - month_date).days + 1)]


def get_months_in_period(start_date, end_date):
    """
    Возвращает список объектов date, представляющих первый день каждого месяца,
    в котором попадает период от start_date до end_date.
    """
    months = []
    current_month = start_date.replace(day=1)
    while current_month <= end_date:
        months.append(current_month)
        if current_month.month == 12:
            current_month = current_month.replace(year=current_month.year + 1, month=1)
        else:
            current_month = current_month.replace(month=current_month.month + 1)
    return months


def get_days_in_period(start_date, end_date):
    """
    Возвращает словарь, где для каждого месяца (ключ в формате 'YYYY-MM')
    указан список дат (объекты date), входящих в выбранный период.
    Если период начинается не с первого дня месяца или заканчивается раньше,
    то в список попадут только дни внутри периода.
    """
    days_by_month = {}
    current = start_date
    while current <= end_date:
        month_key = current.strftime('%Y-%m')
        if month_key not in days_by_month:
            days_by_month[month_key] = {
                'month_name': current.strftime('%B'),
                'days': []
            }
        days_by_month[month_key]['days'].append(current)
        current += timedelta(days=1)
    return days_by_month


@admin_required(login_url='login')
def final_report(request):
    # Получаем даты из GET-параметров
    start_date = request.GET.get('start_date', timezone.now().replace(day=1).date())
    end_date = request.GET.get('end_date', timezone.now().date())

    # Преобразуем даты, если они переданы как строки
    start_date = parse_custom_date(start_date) if isinstance(start_date, str) else start_date
    end_date = parse_custom_date(end_date) if isinstance(end_date, str) else end_date

    # Фильтры (проект, исполнитель, зиС, марка) применяются как в вашем коде…
    selected_projects = request.GET.getlist('project')
    selected_users = request.GET.getlist('employees')
    selected_buildings = request.GET.getlist('zis')
    selected_marks = request.GET.getlist('mark')

    timelogs = Timelog.objects.filter(date__range=[start_date, end_date])

    if selected_projects:
        selected_projects = selected_projects[0].split(',')
        timelogs = timelogs.filter(project__title__in=selected_projects)
    if selected_users:
        user_filters = Q()
        user_list = selected_users[0].split(',')
        for full_name in user_list:
            first_name, last_name = full_name.strip().split(' ', 1)
            user_filters |= Q(user__first_name=first_name, user__last_name=last_name)
        timelogs = timelogs.filter(user_filters)
    if selected_buildings:
        selected_buildings = selected_buildings[0].split(',')
        timelogs = timelogs.filter(building__title__in=selected_buildings)
    if selected_marks:
        selected_marks = selected_marks[0].split(',')
        timelogs = timelogs.filter(mark__title__in=selected_marks)

    timelogs = timelogs.select_related('project', 'building', 'mark', 'user')

    # Сгруппируем логи по уникальному ряду и по дате.
    grouped_data = {}
    for log in timelogs:
        # Формируем уникальный ключ строки. Разделитель можно выбрать любой.
        row_key = f"{log.project.title}|{log.building.title}|{log.mark.title}|{log.user.first_name} {log.user.last_name}"
        date_key = log.date.strftime('%Y-%m-%d')
        if row_key not in grouped_data:
            grouped_data[row_key] = {
                'project': log.project.title,
                'building': log.building.title,
                'mark': log.mark.title,
                'user': f"{log.user.last_name}",
                'total_hours': 0,  # можно суммировать часы по всем датам
                'logs': {}  # здесь будут логи по датам
            }
        grouped_data[row_key]['total_hours'] += log.time
        # Если по одной дате может быть несколько записей – можно делать список.
        grouped_data[row_key]['logs'][date_key] = log.mark.title

    # Если вам требуется оставить и report_data для других целей, можно его оставить:
    report_data = timelogs.values(
        'project__title',
        'building__title',
        'mark__title'
    ).annotate(
        total_hours=Sum('time'),
        user_full_name=Concat(F('user__last_name'), Value(' '), F('user__first_name'))
    )

    # Получаем структуру дней, входящих в выбранный период
    days_by_period = get_days_in_period(start_date, end_date)

    context = {
        'report_data': report_data,  # если нужен для сортировки или прочего
        'grouped_data': grouped_data,  # сгруппированные данные для таблицы
        'days_by_period': days_by_period,  # структура дней по месяцам
        'start_date': start_date,
        'end_date': end_date,
    }

    return render(request, 'task_manager/final_report.html', context)


def parse_russian_date(date_str):
    # Словарь для замены русских названий месяцев на числовые значения
    month_translation = {
        'января': '01', 'февраля': '02', 'марта': '03',
        'апреля': '04', 'мая': '05', 'июня': '06',
        'июля': '07', 'августа': '08', 'сентября': '09',
        'октября': '10', 'ноября': '11', 'декабря': '12'
    }

    # Убираем символ "г." в конце строки
    date_str = re.sub(r'\sг\.$', '', date_str)

    # Заменяем русские названия месяцев на числовые значения
    for ru_month, num_month in month_translation.items():
        date_str = date_str.replace(ru_month, num_month)

    # Преобразуем строку в формат YYYY-MM-DD
    try:
        return datetime.strptime(date_str, '%d %m %Y').date()
    except ValueError as e:
        print(f"Ошибка преобразования даты: {e}")
        return None


def export_to_excel(request):
    # Маппинг английских месяцев на русский
    months_map = {
        "January": "Январь", "February": "Февраль", "March": "Март", "April": "Апрель",
        "May": "Май", "June": "Июнь", "July": "Июль", "August": "Август",
        "September": "Сентябрь", "October": "Октябрь", "November": "Ноябрь", "December": "Декабрь"
    }

    # Получаем строки дат из запроса
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    # Парсим даты с использованием dateparser
    start_date = dateparser.parse(start_date_str).date() if start_date_str else timezone.now().replace(day=1).date()
    end_date = dateparser.parse(end_date_str).date() if end_date_str else timezone.now().date()

    # Проверка на наличие start_date и end_date после парсинга
    if start_date is None or end_date is None:
        print("Error: start_date or end_date is None")
        return HttpResponse("Ошибка: некорректные даты", status=400)

    # Получаем фильтры из GET-запроса
    selected_projects = request.GET.getlist('project')
    selected_users = request.GET.getlist('employees')
    selected_buildings = request.GET.getlist('zis')
    selected_marks = request.GET.getlist('mark')

    # Получаем данные из базы данных
    timelogs = Timelog.objects.filter(date__range=[start_date, end_date])

    # Применяем фильтры, если они указаны
    if selected_projects:
        selected_projects = selected_projects[0].split(',')
        timelogs = timelogs.filter(project__title__in=selected_projects)
    if selected_users:
        user_filters = Q()
        user_list = selected_users[0].split(',')
        for full_name in user_list:
            first_name, last_name = full_name.strip().split(' ', 1)
            user_filters |= Q(user__first_name=first_name, user__last_name=last_name)
        timelogs = timelogs.filter(user_filters)
    if selected_buildings:
        selected_buildings = selected_buildings[0].split(',')
        timelogs = timelogs.filter(building__title__in=selected_buildings)
    if selected_marks:
        selected_marks = selected_marks[0].split(',')
        timelogs = timelogs.filter(mark__title__in=selected_marks)

    # Создаем переменные для хранения марок по дням
    daily_marks = {}
    for log in timelogs:
        # Группировка по: проект, здание, марка и уникальный идентификатор пользователя
        key_str = f"{log.project.title}|{log.building.title}|{log.mark.title}|{log.user.id}"
        day_key = log.date.strftime('%Y-%m-%d')
        if key_str not in daily_marks:
            daily_marks[key_str] = {}
        daily_marks[key_str][day_key] = log.mark.title

    # Создаем новую книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Календарный план"

    # Записываем заголовки
    headers = ["Объект", "ЗиС", "Марка", "Исполнитель", "Часы"]
    days_in_range = pd.date_range(start=start_date, end=end_date)
    headers.extend([day.strftime('%d') for day in days_in_range])  # Только число без месяца
    ws.append(headers)

    # Формируем строку с месяцами
    month_row = ["", "", "", "", ""]
    current_month = ""
    for day in days_in_range:
        month_name = months_map[day.strftime('%B')]
        if month_name != current_month:
            month_row.append(month_name)
            current_month = month_name
        else:
            month_row.append("")
    ws.append(month_row)

    # Обработка данных для экспорта
    for item in timelogs.values(
            'project__title',
            'building__title',
            'mark__title',
            'user__id',
            'user__middle_name'
    ).annotate(
        total_hours=Sum('time'),
        user_full_name=Concat(
            F('user__last_name'),
            Value(' '),
            F('user__first_name'),
            Value(' '),
            F('user__middle_name')
        )
    ):
        # Формируем строку с данными
        row = [
            item['project__title'],
            item['building__title'],
            item['mark__title'],
            item['user_full_name'],  # Теперь включает middle name
            f"{item['total_hours']}"
        ]

        # Формируем группирующий ключ, используя user__id
        key_str = f"{item['project__title']}|{item['building__title']}|{item['mark__title']}|{item['user__id']}"

        for day in days_in_range:
            day_key = day.strftime('%Y-%m-%d')
            mark = daily_marks.get(key_str, {}).get(day_key, "")
            row.append(mark)
        ws.append(row)

    # Настройка ширины колонок
    column_widths = [30, 20, 20, 25, 15] + [5] * len(days_in_range)
    for i, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = column_widths[i - 1]

    # Фиксация заголовков и первых столбцов
    ws.freeze_panes = "F3"  # Фиксируем первую строку с заголовками и первые 5 столбцов

    # Форматирование заголовков
    for row in ws.iter_rows(min_row=1, max_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

    # Выделение суббот и воскресений красным
    for col_num, day in enumerate(days_in_range, start=6):
        if day.weekday() in [5, 6]:  # 5 - суббота, 6 - воскресенье
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_num, max_col=col_num):
                for cell in row:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Добавление границ для всех ячеек
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                 bottom=Side(style='thin'))

    # Генерация файла
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="calendar_plan.xlsx"'
    wb.save(response)
    return response


@login_required
def get_buildings_for_project(request, project_id):
    # Получаем здания для выбранного проекта
    buildings = Building.objects.filter(project_id=project_id)

    # Формируем ответ в виде списка зданий
    building_data = [{'id': building.id, 'name': building.name} for building in buildings]

    return JsonResponse({'buildings': building_data})


def export_reports_employees_excel(request):
    # Получаем начальную и конечную даты из GET параметров
    start_date = request.GET.get('start_date', timezone.now().replace(day=1).date())
    end_date = request.GET.get('end_date', timezone.now().date())

    # Преобразуем даты, если они переданы как строки
    start_date = parse_custom_date(start_date) if isinstance(start_date, str) else start_date
    end_date = parse_custom_date(end_date) if isinstance(end_date, str) else end_date

    # Фильтры из GET-запроса
    selected_projects = request.GET.getlist('project')
    selected_departments = request.GET.getlist('department')
    selected_users = request.GET.getlist('employees')
    selected_stages = request.GET.getlist('stage')
    selected_buildings = request.GET.getlist('zis')
    selected_marks = request.GET.getlist('mark')
    selected_tasks = request.GET.getlist('task')

    # Фильтрация Timelog по диапазону дат с оптимизацией связанных данных
    timelogs = (
        Timelog.objects
        .filter(date__range=[start_date, end_date])
        .select_related('project', 'department', 'user', 'building', 'mark', 'task')
    )

    # Применяем фильтры, если они указаны
    if selected_projects:
        selected_projects = selected_projects[0].split(',')
        timelogs = timelogs.filter(project__title__in=selected_projects)
    if selected_departments:
        selected_departments = selected_departments[0].split(',')
        timelogs = timelogs.filter(department__title__in=selected_departments)
    if selected_users:
        user_filters = Q()
        user_list = selected_users[0].split(',')
        for full_name in user_list:
            first_name, last_name = full_name.strip().split(' ', 1)
            user_filters |= Q(user__first_name=first_name, user__last_name=last_name)
        timelogs = timelogs.filter(user_filters)
    if selected_stages:
        selected_stages = selected_stages[0].split(',')
        timelogs = timelogs.filter(stage__in=selected_stages)
    if selected_buildings:
        selected_buildings = selected_buildings[0].split(',')
        timelogs = timelogs.filter(building__title__in=selected_buildings)
    if selected_marks:
        selected_marks = selected_marks[0].split(',')
        timelogs = timelogs.filter(mark__title__in=selected_marks)
    if selected_tasks:
        selected_tasks = selected_tasks[0].split(',')
        timelogs = timelogs.filter(task__title__in=selected_tasks)

    # Группировка данных по отделам (вместо проектов)
    detailed_report_departments = {}
    for item in timelogs:
        department_title = item.department.title
        if department_title not in detailed_report_departments:
            detailed_report_departments[department_title] = {
                'entries': [],
                'total_time': 0
            }
        detailed_report_departments[department_title]['entries'].append(item)
        detailed_report_departments[department_title]['total_time'] += item.time

    overall_total_time_departments = sum(group['total_time'] for group in detailed_report_departments.values())

    # Создаем книгу Excel
    wb = Workbook()

    # Лист "Отделы"
    ws_departments = wb.active
    ws_departments.title = "Отделы"
    headers_departments = ["Отдел", "Проект", "Дата", "Исполнитель", "Здание", "Марка", "Задача", "Время"]
    ws_departments.append(headers_departments)
    for cell in ws_departments[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Сортируем записи по дате для отделов
    for department_title, group in detailed_report_departments.items():
        entries = sorted(group['entries'], key=lambda x: x.date)
        for entry in entries:
            full_name = f"{entry.user.last_name} {entry.user.first_name}"
            if hasattr(entry.user, 'middle_name') and entry.user.middle_name:
                full_name += f" {entry.user.middle_name}"
            row = [
                entry.department.title,
                entry.project.title,
                entry.date.strftime("%d.%m.%Y"),  # Форматируем дату
                full_name,
                entry.building.title if entry.building else "",
                entry.mark.title if entry.mark else "",
                entry.task.title if entry.task else "",
                entry.time,
            ]
            ws_departments.append(row)
        ws_departments.append(["", "", "", "", "", "Итого по отделу:", group['total_time']])
        ws_departments.append([])

    # Лист "Итого" – сводные итоги
    ws_summary = wb.create_sheet(title="Итого")
    ws_summary.append(["Общая сумма времени по отделам", overall_total_time_departments])
    for cell in ws_summary[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Генерация ответа с файлом Excel
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"reports_employees_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def export_reports_view_excel(request):
    # Получаем начальную и конечную даты из GET параметров
    start_date = request.GET.get('start_date', timezone.now().replace(day=1).date())
    end_date = request.GET.get('end_date', timezone.now().date())

    # Преобразуем даты, если они переданы как строки
    start_date = parse_custom_date(start_date) if isinstance(start_date, str) else start_date
    end_date = parse_custom_date(end_date) if isinstance(end_date, str) else end_date

    # Фильтры из GET-запроса
    selected_projects = request.GET.getlist('project')
    selected_departments = request.GET.getlist('department')
    selected_users = request.GET.getlist('employees')
    selected_stages = request.GET.getlist('stage')
    selected_buildings = request.GET.getlist('zis')
    selected_marks = request.GET.getlist('mark')
    selected_tasks = request.GET.getlist('task')

    # Фильтрация Timelog по диапазону дат с оптимизацией связанных данных
    timelogs = (
        Timelog.objects
        .filter(date__range=[start_date, end_date])
        .select_related('project', 'department', 'user', 'building', 'mark', 'task')
    )

    # Применяем фильтры, если они указаны
    if selected_projects:
        selected_projects = selected_projects[0].split(',')
        timelogs = timelogs.filter(project__title__in=selected_projects)
    if selected_departments:
        selected_departments = selected_departments[0].split(',')
        timelogs = timelogs.filter(department__title__in=selected_departments)
    if selected_users:
        user_filters = Q()
        user_list = selected_users[0].split(',')
        for full_name in user_list:
            first_name, last_name = full_name.strip().split(' ', 1)
            user_filters |= Q(user__first_name=first_name, user__last_name=last_name)
        timelogs = timelogs.filter(user_filters)
    if selected_stages:
        selected_stages = selected_stages[0].split(',')
        timelogs = timelogs.filter(stage__in=selected_stages)
    if selected_buildings:
        selected_buildings = selected_buildings[0].split(',')
        timelogs = timelogs.filter(building__title__in=selected_buildings)
    if selected_marks:
        selected_marks = selected_marks[0].split(',')
        timelogs = timelogs.filter(mark__title__in=selected_marks)
    if selected_tasks:
        selected_tasks = selected_tasks[0].split(',')
        timelogs = timelogs.filter(task__title__in=selected_tasks)

    # Группировка данных по проектам
    detailed_report_projects = {}
    for item in timelogs:
        project_title = item.project.title
        if project_title not in detailed_report_projects:
            detailed_report_projects[project_title] = {
                'entries': [],
                'total_time': 0
            }
        detailed_report_projects[project_title]['entries'].append(item)
        detailed_report_projects[project_title]['total_time'] += item.time

    overall_total_time_projects = sum(group['total_time'] for group in detailed_report_projects.values())

    # Группировка данных по отделам
    detailed_report_departments = {}
    for item in timelogs:
        department_title = item.department.title
        if department_title not in detailed_report_departments:
            detailed_report_departments[department_title] = {
                'entries': [],
                'total_time': 0
            }
        detailed_report_departments[department_title]['entries'].append(item)
        detailed_report_departments[department_title]['total_time'] += item.time

    overall_total_time_departments = sum(group['total_time'] for group in detailed_report_departments.values())

    # Создаем книгу Excel
    wb = Workbook()

    # Лист "Проекты"
    ws_projects = wb.active
    ws_projects.title = "Проекты"
    headers_projects = ["Проект", "Дата", "Отдел", "Исполнитель", "Здание", "Марка", "Задача", "Время"]
    ws_projects.append(headers_projects)
    for cell in ws_projects[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Заполняем данные для проектов. Сортируем записи по дате.
    for project_title, group in detailed_report_projects.items():
        entries = sorted(group['entries'], key=lambda x: x.date)
        for entry in entries:
            # Формируем полное ФИО (добавляем middle name, если есть)
            full_name = f"{entry.user.first_name} {entry.user.last_name}"
            if hasattr(entry.user, 'middle_name') and entry.user.middle_name:
                full_name += f" {entry.user.middle_name}"
            row = [
                entry.project.title,
                entry.date.strftime("%d.%m.%Y"),
                entry.department.title,
                full_name,
                entry.building.title if entry.building else "",
                entry.mark.title if entry.mark else "",
                entry.task.title if entry.task else "",
                entry.time,
            ]
            ws_projects.append(row)
        # Итоговая строка по проекту
        ws_projects.append(["", "", "", "", "", "Итого по проекту:", group['total_time']])
        ws_projects.append([])  # пустая строка для разделения

    # Лист "Отделы"
    ws_departments = wb.create_sheet(title="Отделы")
    headers_departments = ["Отдел", "Проект", "Дата", "Исполнитель", "Здание", "Марка", "Задача", "Время"]
    ws_departments.append(headers_departments)
    for cell in ws_departments[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for department_title, group in detailed_report_departments.items():
        entries = sorted(group['entries'], key=lambda x: x.date)
        for entry in entries:
            full_name = f"{entry.user.last_name} {entry.user.first_name}"
            if hasattr(entry.user, 'middle_name') and entry.user.middle_name:
                full_name += f" {entry.user.middle_name}"
            row = [
                entry.department.title,
                entry.project.title,
                entry.date.strftime("%d.%m.%Y"),
                full_name,
                entry.building.title if entry.building else "",
                entry.mark.title if entry.mark else "",
                entry.task.title if entry.task else "",
                entry.time,
            ]
            ws_departments.append(row)
        ws_departments.append(["", "", "", "", "", "Итого по отделу:", group['total_time']])
        ws_departments.append([])

    # Лист "Итого" – сводные итоги
    ws_summary = wb.create_sheet(title="Итого")
    ws_summary.append(["Общая сумма времени по проектам", overall_total_time_projects])
    ws_summary.append(["Общая сумма времени по отделам", overall_total_time_departments])
    for cell in ws_summary[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Генерация ответа с файлом Excel
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"reports_view_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
